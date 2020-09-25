import openpyxl as xl    #here xl is a alias name to make a code short and more readable
from openpyxl.chart import BarChart,Reference


wb=xl.load_workbook('transactions.xlsx')
sheet= wb['Sheet1']

#cell=sheet['a1']
#cell=sheet.cell(1,1) #// to select cell value
#print(cell.value)  #// to proint that cell value
print('value in third column of Excel sheet are as below: ')
sheet.cell(1,4,'correct value')
for row in range(2, sheet.max_row + 1):
    cell=sheet.cell(row,3)
    print(cell.value)
    corrected_price=cell.value * 0.9
    corrected_price_cell=sheet.cell(row,4)
    corrected_price_cell.value=corrected_price

values=Reference(sheet,
        min_row=2,
        max_row=sheet.max_row,
        min_col=4,
        max_col=4)

chart=BarChart()
chart.add_data(values)
sheet.add_chart(chart,'b7')


wb.save('transaction2.xlsx')
print('')
print("we have to correct tht values \n we have to decrease by 10%  ")

for row in range(2,sheet.max_row+1):
    correct_cell=sheet.cell(row,4)
    print(correct_cell.value)

